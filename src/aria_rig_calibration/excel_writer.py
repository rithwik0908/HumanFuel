"""Excel workbook writer and validator (openpyxl).

Role in the pipeline
--------------------
Write a formatted multi-sheet workbook summarising the run, then validate it structurally against the
source DataFrames. Validation checks sheet presence, sheet order, row/column counts, header names and
order, key totals for the selected-windows sheet, and the absence of forbidden personal columns.
It reports structural equivalence, not a full cell-by-cell value comparison (documented in the report).
No personal identifiers are written.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _safe(name: str) -> str:
    """Return an Excel-legal sheet name (<=31 chars, no ``\\/?*[]:``)."""
    for ch in "\\/?*[]:":
        name = name.replace(ch, " ")
    return name[:31]


def write_workbook(sheet_map: dict[str, pd.DataFrame | None], dest: Path, log) -> Path | None:
    """Write each named DataFrame to its own sheet with bold frozen headers, filters, and widths.

    :param sheet_map: ordered mapping of sheet name -> DataFrame (or None/empty -> a 'no rows' sheet).
    :param dest: output ``.xlsx`` path.
    :param log: collecting logger.
    :return: ``dest`` on success.
    """
    wb = Workbook()
    wb.remove(wb.active)
    for name, df in sheet_map.items():
        ws = wb.create_sheet(_safe(name))
        if df is None or len(df) == 0:
            ws["A1"] = "no rows"
            continue
        ws.append(list(df.columns))
        for _, r in df.iterrows():
            ws.append(["" if pd.isna(v) else v for v in r.tolist()])
        for c in range(1, len(df.columns) + 1):
            ws.cell(1, c).font = Font(bold=True)
            ws.cell(1, c).fill = PatternFill("solid", fgColor="DDDDDD")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for i, col in enumerate(df.columns, start=1):
            mx = df[col].astype(str).str.len().max()
            mx = 8 if pd.isna(mx) else int(mx)
            ws.column_dimensions[get_column_letter(i)].width = min(max(len(str(col)) + 2, mx + 2), 55)
            if any(t in str(col).lower() for t in ("path", "reason", "note", "status", "source")):
                for row in range(2, len(df) + 2):
                    ws.cell(row, i).alignment = Alignment(wrap_text=True)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    log.info("excel: workbook -> %s", dest)
    return dest


def validate_workbook(dest: Path, sheet_map: dict, out_csv: Path, log,
                      forbidden_tokens: list[str] | None = None,
                      report_md: Path | None = None) -> pd.DataFrame:
    """Validate a written workbook structurally against its source frames.

    Per sheet the status is one of ``pass``, ``missing_sheet``, ``row_mismatch``, ``column_mismatch``,
    ``header_mismatch``, ``order_mismatch``, ``key_total_mismatch``, ``privacy_violation``,
    ``read_error``, or ``empty_ok`` (an intentionally empty source sheet). A non-empty sheet whose
    source frame is empty fails as ``row_mismatch`` (empty sheets do not automatically pass).

    :param dest: workbook path.
    :param sheet_map: the same ordered mapping used to write the workbook.
    :param out_csv: where to write the per-sheet validation table.
    :param log: collecting logger.
    :param forbidden_tokens: personal-data tokens that must not appear as column headers.
    :param report_md: optional path for a human-readable summary.
    :return: the validation DataFrame.
    """
    from .metadata import SAFE_ANALYTIC_COLUMNS
    forbidden_tokens = [t.lower() for t in (forbidden_tokens or [])]
    try:
        wb = load_workbook(dest, read_only=True)
    except Exception as e:  # noqa: BLE001
        v = pd.DataFrame([{"sheet": "<workbook>", "status": "read_error", "detail": str(e)}])
        v.to_csv(out_csv, index=False)
        return v

    expected_order = [_safe(n) for n in sheet_map]
    actual_order = list(wb.sheetnames)
    rows = []
    for pos, (name, df) in enumerate(sheet_map.items()):
        sn = _safe(name)
        src_rows = 0 if df is None else len(df)
        src_cols = [] if df is None else [str(c) for c in df.columns]
        detail, status = "", "pass"
        if sn not in wb.sheetnames:
            status, detail = "missing_sheet", "sheet absent from workbook"
        else:
            ws = wb[sn]
            xrows = (ws.max_row - 1) if ws.max_row else 0
            first_row = [c.value for c in next(ws.iter_rows(max_row=1), [])]
            headers = [] if first_row == ["no rows"] else [str(h) for h in first_row if h is not None]
            leaked = [h for h in headers if h.lower() not in SAFE_ANALYTIC_COLUMNS and any(t in h.lower() for t in forbidden_tokens)]
            if src_rows == 0:
                # An expected-empty sheet must contain exactly the "no rows" placeholder in A1.
                if (ws.max_row or 0) > 1:
                    status, detail = "row_mismatch", "empty source but workbook has data rows"
                elif (ws.max_column or 0) > 1:
                    status, detail = "column_mismatch", "empty source but workbook has extra columns"
                elif first_row != ["no rows"]:
                    status, detail = "unexpected_content", f"expected placeholder, found {first_row}"
                else:
                    status = "empty_ok"
            elif leaked:
                status, detail = "privacy_violation", f"forbidden columns: {leaked}"
            elif xrows != src_rows:
                status, detail = "row_mismatch", f"{xrows} vs {src_rows}"
            elif len(headers) != len(src_cols):
                status, detail = "column_mismatch", f"{len(headers)} vs {len(src_cols)}"
            elif headers != src_cols:
                status, detail = "header_mismatch", "header names/order differ"
            elif actual_order[pos] != sn:
                status, detail = "order_mismatch", f"expected pos {pos}"
            # Value spot-check: the Selected Windows sheet's numeric key totals must match the source.
            if status == "pass" and name == "Selected Windows" and df is not None and "score" in df.columns:
                ws_scores = [ws.cell(r, list(src_cols).index("score") + 1).value for r in range(2, xrows + 2)]
                src_total = round(float(pd.to_numeric(df["score"], errors="coerce").sum()), 3)
                xls_total = round(float(pd.Series(ws_scores).astype(float).sum()), 3)
                if abs(src_total - xls_total) > 0.01:
                    status, detail = "key_total_mismatch", f"score total {xls_total} vs {src_total}"
        rows.append({"sheet": name, "expected_pos": pos, "csv_rows": src_rows,
                     "csv_cols": len(src_cols), "status": status, "detail": detail})
    wb.close()  # release the file handle (read_only mode keeps it open) so the run dir can be deleted
    v = pd.DataFrame(rows)
    v.to_csv(out_csv, index=False)
    n_ok = int(v.status.isin(["pass", "empty_ok"]).sum())
    log.info("excel validation: %d/%d sheets ok", n_ok, len(v))
    if report_md is not None:
        bad = v[~v.status.isin(["pass", "empty_ok"])]
        lines = ["# Excel validation", "",
                 f"Structural check (sheet presence/order, row & column counts, header names & "
                 f"order, forbidden-column scan, and the Selected Windows score total). This verifies "
                 f"structure and key totals, not every cell value.", "",
                 f"- Sheets ok: **{n_ok}/{len(v)}**"]
        if len(bad):
            lines.append("- Issues:")
            lines += [f"  - {r.sheet}: {r.status} ({r.detail})" for r in bad.itertuples()]
        else:
            lines.append("- No issues.")
        report_md.write_text("\n".join(lines), encoding="utf-8")
    return v
